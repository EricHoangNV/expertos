#!/usr/bin/env python3
"""Extract the UAT benchmark spreadsheet into two frozen JSON datasets (EN + VI).

The source is a manually-maintained spreadsheet; this step runs once (or whenever
the sheet changes) and writes committed JSON so every benchmark run scores against
an identical, version-controlled input. The runner/scorer never touch the xlsx.

Usage:
    python3 scripts/benchmark/extract.py [path/to/workbook.xlsx]

Output:
    scripts/benchmark/data/dataset.en.json   (100 records, English Q/A)
    scripts/benchmark/data/dataset.vi.json   (100 records, Vietnamese Q/A)

Each record:
    { "id": 1, "lang": "en", "question": "...", "gold_answer": "...",
      "category": "...", "level": "...", "persona": "...",
      "difficulty": "...", "trap": false }
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = ROOT / "tmp" / "Benchmark UAT Questions Design - John Ngo_260711.xlsx"
OUT_DIR = Path(__file__).resolve().parent / "data"

# Sheet layout (row 2 headers, data rows 3..102):
#   A STT | B Câu hỏi (VN) | C Câu trả lời (VN) | D Câu hỏi (EN) | E Câu trả lời (EN) | F GOLD | G Notes
COL_STT, COL_Q_VI, COL_A_VI, COL_Q_EN, COL_A_EN, COL_GOLD = 1, 2, 3, 4, 5, 6
FIRST_ROW = 3


def _clean(v) -> str:
    return "" if v is None else str(v).strip()


def parse_gold(raw) -> dict:
    """Pull Category/Level/Persona/Difficulty/Trap out of the GOLD cell.

    Two formats appear in the sheet: a markdown table (`| **Field** | value |`)
    and a looser alternating-line layout. Best-effort — metadata only drives
    score slicing, so a miss degrades to an empty string, never a crash.
    """
    meta = {"gold_id": "", "category": "", "level": "", "persona": "", "difficulty": "", "trap": False}
    if not raw:
        return meta
    text = str(raw)

    m = re.search(r"GOLD\s+QUESTION\s+([0-9]+)", text, re.IGNORECASE)
    if m:
        meta["gold_id"] = f"GOLD {m.group(1)}"

    fields = {}
    # Markdown-table rows: | **Category** | Strategy ... |
    for row in re.findall(r"\|([^|]+)\|([^|]+)\|", text):
        key = row[0].replace("*", "").strip().lower()
        val = row[1].strip()
        if key and val and not set(val) <= {"-", " "}:
            fields[key] = val
    # Fallback: "Category\n\nValue" alternating lines.
    if not fields:
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        for i, line in enumerate(lines[:-1]):
            k = line.lower()
            if k in ("category", "level", "persona", "difficulty", "trap question"):
                fields[k] = lines[i + 1].strip()

    meta["category"] = fields.get("category", "")
    meta["level"] = fields.get("level", "")
    meta["persona"] = fields.get("persona", "")
    meta["difficulty"] = fields.get("difficulty", "")
    trap = fields.get("trap question", "").lower()
    meta["trap"] = bool(trap) and not trap.startswith("no")
    return meta


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"ERROR: workbook not found: {xlsx}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.worksheets[0]

    en, vi, skipped = [], [], 0
    for r in range(FIRST_ROW, ws.max_row + 1):
        q_en, a_en = _clean(ws.cell(r, COL_Q_EN).value), _clean(ws.cell(r, COL_A_EN).value)
        q_vi, a_vi = _clean(ws.cell(r, COL_Q_VI).value), _clean(ws.cell(r, COL_A_VI).value)
        if not (q_en or q_vi):
            continue
        stt = ws.cell(r, COL_STT).value
        try:
            qid = int(float(stt))
        except (TypeError, ValueError):
            qid = len(en) + 1
        meta = parse_gold(ws.cell(r, COL_GOLD).value)

        if q_en and a_en:
            en.append({"id": qid, "lang": "en", "question": q_en, "gold_answer": a_en, **{k: v for k, v in meta.items() if k != "gold_id"}, "gold_id": meta["gold_id"]})
        else:
            skipped += 1
        if q_vi and a_vi:
            vi.append({"id": qid, "lang": "vi", "question": q_vi, "gold_answer": a_vi, **{k: v for k, v in meta.items() if k != "gold_id"}, "gold_id": meta["gold_id"]})
        else:
            skipped += 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "dataset.en.json").write_text(json.dumps(en, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "dataset.vi.json").write_text(json.dumps(vi, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Source: {xlsx}")
    print(f"  EN: {len(en)} records -> {OUT_DIR / 'dataset.en.json'}")
    print(f"  VI: {len(vi)} records -> {OUT_DIR / 'dataset.vi.json'}")
    if skipped:
        print(f"  (skipped {skipped} incomplete half-records)")
    cats = sorted({rec['category'] for rec in en if rec['category']})
    print(f"  distinct categories: {len(cats)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
